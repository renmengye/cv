#!/usr/bin/env python3
"""
Compile an NSF Collaborators and Other Affiliations (CoA) xlsx from a unified
YAML source of truth (coa.yaml) plus a publications.yaml.

    coa.yaml   ── advisors, advisees, coauthor institutions, other collabs
    publications.yaml ── source of author names (last 48 months)
    template.xlsx ── NSF CoA template (blank; used for formatting)
         │
         ▼
    coa_output.xlsx

The script:
    - Reads coa.yaml (single source of truth for institutions/relationships).
    - Reads publications.yaml, filters to the 48-month window ending at
      --submission-date, extracts unique co-authors.
    - For each co-author, looks up institution in coa.yaml's
      coauthor_institutions map. Rows with blank institution are still
      written to the xlsx (so you see who is missing) and printed as an
      "unresolved" summary at the end.
    - Writes Tables 1, 3, 4, 5 into the NSF template; inserts blank rows
      as needed and clones formatting from example template rows.

Usage:
    python3 build_coa.py \\
        --coa nsf-career-2026/coa.yaml \\
        --publications ~/code/cv/publications.yaml \\
        --template "/Users/mengye/coa_template (15).xlsx" \\
        --output nsf-career-2026/coa_ren_2026_07.xlsx \\
        --submission-date 2026-07-17
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from copy import copy
from pathlib import Path

import openpyxl
import yaml

WINDOW_MONTHS = 48

TEMPLATE = {
    "table1_example_row": 17,
    "table3_example_row": 38,
    "table3_first_data":  38,
    "table4_example_row": 52,
    "table4_first_data":  52,
    "table5_example_row": 64,
    "table5_first_data":  64,
}

# ---------------------------------------------------------------------------
# Name helpers
# ---------------------------------------------------------------------------
def normalize_name(raw: str) -> tuple[str, str]:
    raw = raw.strip().rstrip(",")
    if "," in raw:
        last, rest = raw.split(",", 1)
        return last.strip(), rest.strip()
    parts = raw.rsplit(" ", 1)
    if len(parts) == 1:
        return parts[0], ""
    return parts[1].strip(), parts[0].strip()


def display_key(last: str, first: str) -> str:
    """coauthor_institutions key: 'last, first_token' (lowercase; first_token
    = first word of given name — so 'Mozer, Michael C.' maps to 'mozer, michael')."""
    first_token = first.strip().split()[0] if first.strip() else ""
    return f"{last.lower().strip()}, {first_token.lower()}"


# ---------------------------------------------------------------------------
# Publications parsing
# ---------------------------------------------------------------------------
def extract_coauthors(pubs_path: Path, submission_date: dt.date,
                     pi_last: str, pi_first: str) -> dict:
    data = yaml.safe_load(pubs_path.read_text())
    pubs = data.get("publications", data if isinstance(data, list) else [])
    window_start = dt.date(submission_date.year - WINDOW_MONTHS // 12,
                           submission_date.month, 1)
    pi_dk = display_key(pi_last, pi_first)

    coauthors: dict = {}
    for pub in pubs:
        year = pub.get("year")
        if not year:
            continue
        # If month is unknown, default to January (conservative — excludes
        # early-year papers that might fall outside the 48-month window).
        try:
            pd = dt.date(int(year), int(pub.get("month") or 1), 1)
        except (ValueError, TypeError):
            continue
        if pd < window_start:
            continue
        for auth in pub.get("authors") or []:
            last, first = normalize_name(auth)
            dk = display_key(last, first)
            if dk == pi_dk:
                continue
            entry = coauthors.setdefault(dk, {
                "last": last, "first": first,
                "max_date": pd, "n_papers": 0,
            })
            if pd > entry["max_date"]:
                entry["max_date"] = pd
            entry["n_papers"] += 1
            # Keep the fuller first-name form seen.
            if len(first) > len(entry["first"]):
                entry["first"] = first
    return coauthors


# ---------------------------------------------------------------------------
# Template writer helpers
# ---------------------------------------------------------------------------
def clone_row_format(ws, src_row: int, dst_row: int, ncols: int = 6) -> None:
    for c in range(1, ncols + 1):
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        if s.has_style:
            d.font = copy(s.font)
            d.alignment = copy(s.alignment)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.number_format = s.number_format


def clear_template_examples(ws) -> None:
    for row_group in [(18, 19), (28,), (38, 39), (52, 53), (64, 65)]:
        for r in row_group:
            for c in range(1, 6):
                ws.cell(row=r, column=c).value = None


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------
def build(args) -> int:
    sub_date = dt.date.fromisoformat(args.submission_date)
    coa = yaml.safe_load(Path(args.coa).read_text())

    pi = coa["pi"]
    advisors = coa.get("advisors", [])
    advisees = coa.get("advisees", [])
    inst_map = coa.get("coauthor_institutions", {}) or {}
    other_collabs = coa.get("other_collaborators", []) or []
    editors = coa.get("editors", []) or []

    print(f"[coa.yaml]      {len(advisors)} advisors, {len(advisees)} advisees, "
          f"{len(inst_map)} coauthor slots, {len(other_collabs)} other collabs, "
          f"{len(editors)} editors")

    # 1. Extract co-authors from publications
    coauthors = extract_coauthors(Path(args.publications), sub_date,
                                  pi["last"], pi["first"])
    print(f"[publications]  {len(coauthors)} unique co-authors in the last {WINDOW_MONTHS} months")

    # 2. Build Table 3 rows.
    # NSF code convention:
    #   G:  people who advised the PI during PI's PhD       (PI's advisors)
    #   T:  people the PI has advised through PhD thesis    (PI's advisees)
    table3: list[list] = []
    for e in advisors:
        table3.append([
            f"{e.get('code', 'G')}:",
            f"{e['last']}, {e['first']}",
            e.get("inst", ""),
            e.get("orcid", ""),
            "",
        ])
    for e in advisees:
        table3.append([
            f"{e.get('code', 'T')}:",
            f"{e['last']}, {e['first']}",
            e.get("inst", ""),
            e.get("orcid", ""),
            "",
        ])

    # 3. Build Table 4 rows: A: co-authors (from publications, sorted by last name),
    #    then C: other_collaborators.
    coauthor_rows: list[list] = []
    unresolved: list[dict] = []
    # Sort co-authors by last active date descending (most recent first),
    # then alphabetically by (last, first) as tiebreaker.
    for dk, ca in sorted(coauthors.items(),
                          key=lambda x: (-x[1]["max_date"].toordinal(),
                                         x[1]["last"].lower(),
                                         x[1]["first"].lower())):
        inst = inst_map.get(dk, "")
        row = [
            "A:",
            f"{ca['last']}, {ca['first']}",
            inst,
            "",
            ca["max_date"].strftime("%m/%d/%Y"),
        ]
        coauthor_rows.append(row)
        if not inst:
            unresolved.append({"key": dk, "n_papers": ca["n_papers"],
                                "max_year": ca["max_date"].year})

    other_rows: list[list] = []
    for e in other_collabs:
        year = e.get("last_active", "")
        year_s = f"01/01/{year}" if year else ""
        other_rows.append([
            "C:",
            f"{e['last']}, {e['first']}",
            e.get("inst", ""),
            e.get("orcid", ""),
            year_s,
        ])

    table4 = coauthor_rows + other_rows

    # 4. Editors -> Table 5
    table5: list[list] = []
    for e in editors:
        table5.append([
            f"{e.get('code', 'B')}:",
            f"{e['last']}, {e['first']}",
            e.get("inst", ""),
            e.get("journal", ""),
            f"01/01/{e['last_active']}" if e.get("last_active") else "",
        ])

    # 5. Write output using template as base
    wb = openpyxl.load_workbook(args.template)
    ws = wb[wb.sheetnames[0]]

    # Remove any Excel Table objects the template ships with — they carry
    # banded-row styling and autofilter dropdowns, and their data ranges
    # do NOT automatically expand when we insert new rows. Leaving them in
    # causes orphan header stripes and inconsistent row shading (e.g. one
    # random row in the middle of Table 4 highlighted blue).
    for tbl_name in list(ws.tables.keys()):
        del ws.tables[tbl_name]

    clear_template_examples(ws)

    # Table 1: PI + affiliation
    ws.cell(row=17, column=2).value = f"{pi['last']}, {pi['first']}"
    ws.cell(row=17, column=3).value = pi["affiliation"]

    # Insert rows in DESCENDING template order so anchors above stay stable.

    # Table 5 (editors) — fixed slot, only insert if we have multiple editors
    n5 = max(1, len(table5))
    if n5 > 1:
        ws.insert_rows(TEMPLATE["table5_first_data"] + 1, amount=n5 - 1)
    for i, cells in enumerate(table5):
        r = TEMPLATE["table5_first_data"] + i
        if i > 0:
            clone_row_format(ws, TEMPLATE["table5_example_row"], r)
        for c, val in enumerate(cells, 1):
            ws.cell(row=r, column=c).value = val

    # Table 4
    n4 = max(1, len(table4))
    if n4 > 1:
        ws.insert_rows(TEMPLATE["table4_first_data"] + 1, amount=n4 - 1)
    for i, cells in enumerate(table4):
        r = TEMPLATE["table4_first_data"] + i
        if i > 0:
            clone_row_format(ws, TEMPLATE["table4_example_row"], r)
        for c, val in enumerate(cells, 1):
            ws.cell(row=r, column=c).value = val

    # Table 3
    n3 = max(1, len(table3))
    if n3 > 1:
        ws.insert_rows(TEMPLATE["table3_first_data"] + 1, amount=n3 - 1)
    for i, cells in enumerate(table3):
        r = TEMPLATE["table3_first_data"] + i
        if i > 0:
            clone_row_format(ws, TEMPLATE["table3_example_row"], r)
        for c, val in enumerate(cells, 1):
            ws.cell(row=r, column=c).value = val

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    # 6. Report
    print()
    print(f"[output] {args.output}")
    print(f"  Table 3: {len(table3)} rows (advisors + advisees)")
    print(f"  Table 4: {len(table4)} rows ({len(coauthor_rows)} A: co-authors, {len(other_rows)} C: other)")
    print(f"  Table 5: {len(table5)} rows (editors)")

    # 7. List unresolved so PI can fill coa.yaml
    if unresolved:
        # Warn about extra keys in coa.yaml not seen in publications (stale)
        stale = sorted(set(inst_map.keys()) - set(coauthors.keys()))
        if stale:
            print()
            print(f"[stale] {len(stale)} entries in coa.yaml's coauthor_institutions "
                  f"do NOT match any co-author in the current publication window:")
            for k in stale[:10]:
                print(f"  - {k!r}  (institution: {inst_map[k]!r})")
            if len(stale) > 10:
                print(f"  ... {len(stale) - 10} more")

        print()
        print(f"[unresolved] {len(unresolved)} co-authors have blank institution.")
        print("Fill the RHS in coa.yaml → coauthor_institutions, then re-run.")
        print()
        # Sort unresolved by n_papers descending
        unresolved.sort(key=lambda x: (-x["n_papers"], x["key"]))
        show = unresolved[:args.batch_size] if args.batch_size > 0 else unresolved
        for u in show:
            print(f"  {u['key']!r}: \"\"    # {u['n_papers']} joint pub(s), last active {u['max_year']}")
        if args.batch_size > 0 and len(unresolved) > args.batch_size:
            print(f"  ... plus {len(unresolved) - args.batch_size} more")
    else:
        print()
        print("[ok] All co-authors have institutions.")
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    p.add_argument("--coa", required=True,
                   help="Path to coa.yaml (single source of truth)")
    p.add_argument("--publications", required=True)
    p.add_argument("--template", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--submission-date", required=True)
    p.add_argument("--batch-size", type=int, default=20,
                   help="How many unresolved co-authors to print at the end (0 = all)")
    args = p.parse_args()
    return build(args)


if __name__ == "__main__":
    sys.exit(main())
